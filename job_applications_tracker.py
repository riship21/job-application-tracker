from openpyxl import load_workbook
from datetime import datetime

# Load the existing workbook
try:
    workbook = load_workbook('job_applications.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = load_workbook()
    sheet = workbook.active
    headings = ["Company Name", "Job Title", "Application Date", "Status"]
    sheet.append(headings)

# Prompt the user for input and store it in the Excel sheet
company_name = input("Enter the company name: ")
job_title = input("Enter the job title: ")

# Get the current date
current_date = datetime.now().strftime('%m-%d-%Y')

status = input("Enter the status of the application: ")

# Append the input data to the Excel sheet
sheet.append([company_name, job_title, current_date, status])

# Save the workbook
workbook.save('job_applications.xlsx')
print("Data saved successfully.")
